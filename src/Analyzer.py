import re
from itertools import count

import openpyxl

RIC = 1
FID = 3
IDN = 4
ERT = 5
FLAG = 6
ASCII_A = 65
MISMATCH = "!"


def checkIfFormatting(IDN_local, ERT_local):
    try:
        ERT_local_var = re.sub('[""]', '', ERT_local)
        IDN_local_var = re.sub('[""]', '', IDN_local)

        if ERT_local_var != 'null' and IDN_local_var != 'null':
            if ERT_local_var != IDN_local_var and float(ERT_local_var) == float(IDN_local_var):
                return True
    except:
        return False

def checkIfMinorValueDifference(IDN_local, ERT_local):
    try:
        IDN_local_var = re.sub('[""]', '', IDN_local)
        ERT_local_var = re.sub('[""]', '', ERT_local)

        if ERT_local_var != 'null' and IDN_local_var != 'null' and len(IDN_local) != 1:
            # print("check" + IDN_local_var + " " + ERT_local_var)
            if IDN_local_var != ERT_local_var and float(IDN_local_var) != float(ERT_local_var):
                # print(IDN_local_var + " " + ERT_local_var)
                tmp = round(float(ERT_local_var) - float(IDN_local_var), 2)
                # print(tmp)
                if tmp == 0.10 or tmp == -0.10:
                    return True
                elif tmp == 0.01 or tmp == -0.01:
                    return True
                elif tmp == 0.001 or tmp == -0.001:
                    return True
    except:
        # print("Exception in calculations.")
        # print(ERT_local_var + " " + IDN_local_var)
        return False

def checkIfSpacingIsAcceptable(IDN_local, ERT_local):
    length_ERT = len(ERT_local)
    length_IDN = len(IDN_local)

    if (length_ERT + IDN_local.count(' ')) == length_IDN and length_IDN != 1 and IDN_local.count(' ') != 0:
        return True
    return False

#TODO make it universal
def checkIfAcceptableMismatch(IDN_local, ERT_local, AcceptableGeneralMismatch):
    ERT_local_var = re.sub('[""]', '', ERT_local)
    IDN_local_var = re.sub('[""]', '', IDN_local)

    if IDN_local_var == AcceptableGeneralMismatch[0][0] and ERT_local_var == \
            AcceptableGeneralMismatch[0][1]:
        return True
    elif IDN_local_var == AcceptableGeneralMismatch[1][0] and ERT_local_var == \
            AcceptableGeneralMismatch[1][1]:
        return True
    return False

def checkConditions(lineToAnalyze, AcceptableGeneralMismatch):
    if checkIfAcceptableMismatch(lineToAnalyze[IDN], lineToAnalyze[ERT], AcceptableGeneralMismatch):
        # print("here1" + lineToAnalyze[FID])
        return True
    elif checkIfFormatting(lineToAnalyze[IDN], lineToAnalyze[ERT]):
        # print("here2" + lineToAnalyze[FID])
        return True
    elif checkIfSpacingIsAcceptable(lineToAnalyze[IDN], lineToAnalyze[ERT]):
        # print("here3" + lineToAnalyze[FID])
        return True
    elif checkIfMinorValueDifference(lineToAnalyze[IDN], lineToAnalyze[ERT]):
        # print("here4" + lineToAnalyze[FID])
        return True
    return False


class Analyzer:
    dictVar = {}
    dictVarQuantityOfMismatches = {}

    def __init__(self, Data_to_analyze, AcceptableGeneralMismatch, FIDsNotToBeAnalyzed, CID):
        self.list_of_fids_with_mismatch = []
        self.filenameToWrite = 'Result_' + CID + '.xlsx'

        for row in Data_to_analyze:
            lineToAnalyze = row.split(",")

            if len(lineToAnalyze) != 1:
                if lineToAnalyze[FLAG] == MISMATCH:
                    if lineToAnalyze[FID] not in FIDsNotToBeAnalyzed:
                        if not checkConditions(lineToAnalyze, AcceptableGeneralMismatch):
                            if lineToAnalyze[FID] not in self.list_of_fids_with_mismatch:
                                self.list_of_fids_with_mismatch.append(lineToAnalyze[FID])
                                self.dictVar[lineToAnalyze[FID]] = []
                                self.dictVarQuantityOfMismatches[lineToAnalyze[FID]] = 0

                            sublist = [lineToAnalyze[RIC], lineToAnalyze[IDN], lineToAnalyze[ERT]]
                            self.dictVar[lineToAnalyze[FID]].append(sublist)
                            self.dictVarQuantityOfMismatches[lineToAnalyze[FID]] = self.dictVarQuantityOfMismatches[lineToAnalyze[FID]] + 1
            else:
                continue

    def columnAutoWidth(self, wbSheet, *val):
        i = ASCII_A
        for n in val:
            col_width = wbSheet.column_dimensions[chr(i)].width
            if n > col_width:
                wbSheet.column_dimensions[chr(i)].width = n
            i += 1

    def generateSheetsResultsForFIDs(self):
        wb = openpyxl.load_workbook(filename=self.filenameToWrite)

        sheetnames = wb.get_sheet_names()
        # print(sheetnames)

        for sheet in sheetnames:
            if sheet != 'Sheet':
                col_width = [0,0,0]
                wbSheet = wb[sheet]
                dataForSheet = self.dictVar[sheet]
                wbSheet.cell(row=1, column=1).value = "RIC"
                wbSheet.cell(row=1, column=2).value = "IDN"
                wbSheet.cell(row=1, column=3).value = "ERT"
                i = 2
                for singleLine in dataForSheet:
                    wbSheet.cell(row=i, column=1).value = singleLine[0]
                    wbSheet.cell(row=i, column=2).value = singleLine[1]
                    wbSheet.cell(row=i, column=3).value = singleLine[2]
                    i = i + 1
                    col_width = [max(col_width[0],len(singleLine[0])),max(col_width[1],len(singleLine[1])),max(col_width[2],len(singleLine[2]))]
                self.columnAutoWidth(wbSheet, col_width[0], col_width[1], col_width[2])
            else:
                col_width = [0, 0]
                wbSheet = wb[sheet]
                wbSheet.cell(row=1, column=1).value = "FID"
                wbSheet.cell(row=1, column=2).value = "Mismatches quantity"
                i = 2
                for MismatchedFID in self.list_of_fids_with_mismatch:
                    wbSheet.cell(row=i, column=1).value = MismatchedFID
                    wbSheet.cell(row=i, column=2).value = self.dictVarQuantityOfMismatches[MismatchedFID]
                    i = i + 1
                    col_width = [max(col_width[0], len(MismatchedFID)), max(col_width[1], len(str(self.dictVarQuantityOfMismatches[MismatchedFID])))]
                self.columnAutoWidth(wbSheet, col_width[0], col_width[1])
                wbSheet.title = "Statistics"
        wb.save(filename=self.filenameToWrite)
        wb.close()

    def getListOfFIDWithMismatch(self):
        ws_name = self.filenameToWrite
        rb = openpyxl.load_workbook(ws_name)

        for MismatchedFID in self.list_of_fids_with_mismatch:
            rb.create_sheet(MismatchedFID)

        rb.save(ws_name)
        rb.close()

        self.generateSheetsResultsForFIDs()

    def getDetailsOnMismatchesForFID(self, SpecifiedFID):
        for MismatchedUpdate in self.dictVar[str(SpecifiedFID)]:
            print(MismatchedUpdate)

    def deleteFIDFromAnalysis(self, FIDVar):
        del (self.dictVar[FIDVar])
        del (self.dictVarQuantityOfMismatches[FIDVar])
        self.list_of_fids_with_mismatch.remove(FIDVar)


class AnalyzerWrapper(object):

    def __init__(self, Data_to_analyze, AcceptableGeneralMismatch, FIDsNotToBeAnalyzed, CID):
        self.Data_to_analyze = Data_to_analyze
        self.AcceptableGeneralMismatch = AcceptableGeneralMismatch
        self.FIDsNotToBeAnalyzed = FIDsNotToBeAnalyzed
        self.AnalyzerVar = Analyzer(Data_to_analyze, AcceptableGeneralMismatch, FIDsNotToBeAnalyzed, CID)

    def execute_option(self, argument):
        method_name = 'option_' + str(argument)
        method = getattr(self, method_name, lambda: "Invalid option")
        return method()

    def option_1(self):
        return self.AnalyzerVar.getListOfFIDWithMismatch()

    def option_2(self):
        try:
            # TODO: list of available options: 1 - FID, different function for convertion
            option = input("Please, enter FID for analysis.\n")
            return self.AnalyzerVar.getDetailsOnMismatchesForFID(option)
        except:
            print("Please, valid FID from options list should be selected.")

    def option_3(self):
        try:
            # TODO: list of available options: 1 - FID, different function for convertion
            option = input("Please, enter FID for deletion from analysis.\n")
            return self.AnalyzerVar.deleteFIDFromAnalysis(option)
        except:
            print("Please, valid FID from options list should be selected.")

    def option_4(self):
        return quit()
