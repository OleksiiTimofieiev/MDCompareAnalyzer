#!/usr/bin/python

import sys
import openpyxl
import math

limit = 40000

class FileReader:
    sheets = []
    row_quantity = 0
    counter_for_progress_bar_limiter = limit

    def __init__(self, filenameExcel, CID):
        try:
            self.wb = openpyxl.load_workbook(filenameExcel, read_only=True)
        except:
            print("No such file or filename contains spaces")
            sys.exit()
        self.filenameNew = 'Result_' + CID + '.xlsx'
        wb = openpyxl.Workbook()
        wb.save(filename=self.filenameNew)
        wb.close()

        for x in self.wb.sheetnames:
            if x.find("Vertical") == 0:
                self.sheets.append(x)
        for x in self.sheets:
            self.row_quantity += self.wb[x].max_row

        self.data_to_analyze = []

        for sheet in self.sheets:
            data = self.wb[sheet].rows

            toolbar_width = math.ceil(self.wb[sheet].max_row / self.counter_for_progress_bar_limiter)

            counter_for_progress_bar = 0

            sys.stdout.write(
                "[%s] rows to be processed => %s" % (" " * (int(toolbar_width)), str(self.wb[sheet].max_row)))
            sys.stdout.flush()
            sys.stdout.write("\b" * (32 + int(toolbar_width)))  # return to start of line, after '['
            test = 0
            for row in data:
                list_row = list(row)
                list_row_str = ""

                for i in range(len(list_row)):
                    if i == len(list_row) - 1:
                        list_row_str += str(list_row[i].value)
                    else:
                        list_row_str += (str(list_row[i].value) + ',')
                self.data_to_analyze.append(list_row_str)

                if counter_for_progress_bar % self.counter_for_progress_bar_limiter == 0:
                    sys.stdout.write("#")
                    sys.stdout.flush()
                # test = test + 1
                # if test == 100000:
                #     break
                counter_for_progress_bar = counter_for_progress_bar + 1
            sys.stdout.write("]\n")  # this ends the progress bar
        self.wb.close()

    def getDataForAnalysis(self):
        return self.data_to_analyze

