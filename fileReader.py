#!/usr/bin/python

from datetime import datetime
import time
# from progress.bar import IncrementalBar
# from progress.bar import Bar
# from alive_progress import alive_bar
# from progress.bar import IncrementalBar

import csv
import sys
import xlrd
import pandas as pd
import openpyxl
import math


# def convertTuple(tup):
#     str = ''.join(tup)
#     return str

# startTime = datetime.now()
class FileReader:
    sheets = []
    row_quantity = 0

    def __init__(self, filenameExcel):
        self.wb = openpyxl.load_workbook(filenameExcel, read_only=True)
        for x in self.wb.sheetnames:
            if x.find("Vertical") == 0:
                self.sheets.append(x)
        for x in self.sheets:
            self.row_quantity += self.wb[x].max_row
        # print(self.row_quantity)

        self.data_to_analyze = []

        # print("Sheets to be processed -> " + str(len(self.sheets)))

        for sheet in self.sheets:
            data = self.wb[sheet].rows

            toolbar_width = math.ceil(self.wb[sheet].max_row / 40000)
            # print(toolbar_width)
            counter_for_progress_bar = 0
            # print(sheet + " rows quantity:\n" +  + " items")
            sys.stdout.write(
                "[%s] rows to be processed => %s" % (" " * (int(toolbar_width)), str(self.wb[sheet].max_row)))
            sys.stdout.flush()
            sys.stdout.write("\b" * (32 + int(toolbar_width)))  # return to start of line, after '['

            for row in data:
                list_row = list(row)
                list_row_str = ""

                for i in range(len(list_row)):
                    if i == len(list_row) - 1:
                        list_row_str += str(list_row[i].value)
                    else:
                        list_row_str += (str(list_row[i].value) + ',')
                self.data_to_analyze.append(list_row_str)

                if counter_for_progress_bar % 40000 == 0:
                    sys.stdout.write("#")
                    sys.stdout.flush()
                counter_for_progress_bar = counter_for_progress_bar + 1
            sys.stdout.write("]\n")  # this ends the progress bar
        self.wb.close()

    def getDataForAnalysis(self):
        return self.data_to_analyze

        # # Opening a file
        # file1 = open('test', 'w')
        #
        # # Writing a string to file
        # for x in z:
        #     file1.write(x + '\n')
        #
        # # Writing multiple strings
        # # at a time
        # # file1.writelines(L)
        #
        # # Closing file
        # file1.close()
        # # print(z)
